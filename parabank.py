from pathlib import Path
import re
import logging
from datetime import datetime
from dataclasses import dataclass
from typing import List
import pandas as pd
from dateutil import parser as date_parser
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule

# PATHS
SCRIPT_DIR = Path(__file__).resolve().parent
CWD = Path.cwd()
CSV_NAME = "ParaBank_users.csv"

if (SCRIPT_DIR / CSV_NAME).exists():
    CSV_PATH = SCRIPT_DIR / CSV_NAME
elif (CWD / CSV_NAME).exists():
    CSV_PATH = CWD / CSV_NAME
else:
    raise FileNotFoundError("CSV NOT FOUND")

REPORT_XLSX = SCRIPT_DIR / "Parabank_Automation_Report.xlsx"
SCREENSHOT_DIR = SCRIPT_DIR / "screenshots"
SCREENSHOT_DIR.mkdir(exist_ok=True)


# CONFIG
BASE_URL = "https://parabank.parasoft.com/parabank/index.htm  "
USD_TO_EUR = 0.92
LOAN_AMOUNT_USD = 10000
DOWNPAYMENT_FACTOR = 0.20
TIMEOUT = 15
DEFAULT_INITIAL_DEPOSIT = 500.0  

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("parabank")


@dataclass
class User:
    first_name: str
    last_name: str
    address: str
    city: str
    state: str
    zip_code: str
    phone: str
    ssn: str
    username: str
    password: str
    dob_normalized: str
    initial_deposit: float
    debit_card: str

def normalize_dob(raw_dob: str) -> str:
    if not raw_dob or raw_dob.strip() == "":
        return ""
    try:
        dt = date_parser.parse(raw_dob, fuzzy=True)
        return f"{dt.month}/{dt.day}/{dt.year}"
    except Exception:
        return raw_dob.strip()

def validate_user(u: User) -> List[str]:
    errors = []
    required = ["first_name", "address", "city", "state", "zip_code", "phone", "ssn", "username", "password"]
    for field in required:
        val = getattr(u, field)
        if not val or str(val).strip() == "":
            errors.append(f"missing {field}")
    if u.zip_code and not re.fullmatch(r"\d{5}", str(u.zip_code)):
        errors.append("zip_code must be 5-digit number")
    return errors


def make_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--window-size=1400,900")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def wait(driver, by, value, timeout=TIMEOUT):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))


def register(driver, u: User):
    driver.find_element(By.LINK_TEXT, "Register").click()
    wait(driver, By.ID, "customer.firstName")
    driver.find_element(By.ID, "customer.firstName").send_keys(u.first_name)
    driver.find_element(By.ID, "customer.lastName").send_keys(u.last_name or "")
    driver.find_element(By.ID, "customer.address.street").send_keys(u.address)
    driver.find_element(By.ID, "customer.address.city").send_keys(u.city)
    driver.find_element(By.ID, "customer.address.state").send_keys(u.state)
    driver.find_element(By.ID, "customer.address.zipCode").send_keys(u.zip_code)
    driver.find_element(By.ID, "customer.phoneNumber").send_keys(u.phone)
    driver.find_element(By.ID, "customer.ssn").send_keys(u.ssn)
    driver.find_element(By.ID, "customer.username").send_keys(u.username)
    driver.find_element(By.ID, "customer.password").send_keys(u.password)
    driver.find_element(By.ID, "repeatedPassword").send_keys(u.password)
    driver.find_element(By.XPATH, "//input[@value='Register']").click()
    wait(driver, By.LINK_TEXT, "Log Out")


def open_account(driver):
    driver.find_element(By.LINK_TEXT, "Open New Account").click()
    wait(driver, By.ID, "type")
    driver.find_element(By.XPATH, "//input[@value='Open New Account']").click()
    try:
        return wait(driver, By.ID, "newAccountId").text
    except TimeoutException:
        return ""

def request_loan(driver, down_payment):
    driver.find_element(By.LINK_TEXT, "Request Loan").click()
    wait(driver, By.ID, "amount")
    driver.find_element(By.ID, "amount").send_keys(str(LOAN_AMOUNT_USD))
    driver.find_element(By.ID, "downPayment").send_keys(str(down_payment))
    driver.find_element(By.XPATH, "//input[@value='Apply Now']").click()
    try:
        status = wait(driver, By.ID, "loanStatus").text
    except TimeoutException:
        status = "UNKNOWN"
    try:
        acc = driver.find_element(By.ID, "newAccountId").text
    except:
        acc = ""
    return status, acc


def format_excel_report(path: Path, df: pd.DataFrame):
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Automation Report"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin

    for col_name in ["Loan USD", "Loan EUR", "Down Payment USD", "Down Payment EUR"]:
        if col_name in df.columns:
            idx = df.columns.get_loc(col_name) + 1
            letter = ws.cell(row=1, column=idx).column_letter
            for row in range(2, ws.max_row + 1):
                ws[f"{letter}{row}"].number_format = "#,##0.00"

    if "Status" in df.columns:
        status_idx = df.columns.get_loc("Status") + 1
        letter = ws.cell(row=1, column=status_idx).column_letter
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            FormulaRule(
                formula=[f'{letter}2="COMPLETED"'],
                fill=PatternFill("solid", fgColor="C6EFCE"),
                font=Font(color="006100")
            )
        )
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            FormulaRule(
                formula=[f'{letter}2="FAILED"'],
                fill=PatternFill("solid", fgColor="FFC7CE"),
                font=Font(color="9C0006")
            )
        )

    wb.save(path)

def main():
    df = pd.read_csv(CSV_PATH, dtype=str)
    df = df.fillna("")
    report = []

    for _, r in df.iterrows():
        # Extract fields exactly as named in CSV
        first_name = str(r["First Name"]).strip()
        last_name = str(r["Last Name"]).strip()
        address = str(r["Address"]).strip()
        city = str(r["City"]).strip()
        state = str(r["State"]).strip()
        zip_code = str(r["Zip Code"]).strip()
        phone = str(r["Phone Number"]).strip()
        ssn = str(r["SSN"]).strip()
        username = str(r["Username"]).strip()
        password = str(r["Password"]).strip()
        dob_raw = str(r["DOB"]).strip() if pd.notna(r["DOB"]) else ""
        debit_card = str(r["Debit Card"]).replace(" ", "") if pd.notna(r["Debit Card"]) else "N/A"

        # Normalize DOB 
        dob_normalized = normalize_dob(dob_raw)

        # Handle Initial Deposit: use default if missing
        deposit_str = str(r["Initial Deposit"]).strip()
        if deposit_str == "" or deposit_str.lower() in ("nan", "null"):
            initial_deposit = DEFAULT_INITIAL_DEPOSIT
        else:
            try:
                initial_deposit = float(deposit_str)
            except (ValueError, TypeError):
                initial_deposit = DEFAULT_INITIAL_DEPOSIT

        user = User(
            first_name=first_name,
            last_name=last_name,
            address=address,
            city=city,
            state=state,
            zip_code=zip_code,
            phone=phone,
            ssn=ssn,
            username=username,
            password=password,
            dob_normalized=dob_normalized,
            initial_deposit=initial_deposit,
            debit_card=debit_card
        )

        errors = validate_user(user)
        down_payment = round(user.initial_deposit * DOWNPAYMENT_FACTOR, 2)

        row = {
            "Username": user.username,
            "DOB": user.dob_normalized,
            "Debit Card": user.debit_card,
            "Loan USD": LOAN_AMOUNT_USD,
            "Loan EUR": round(LOAN_AMOUNT_USD * USD_TO_EUR, 2),
            "Down Payment USD": down_payment,
            "Down Payment EUR": round(down_payment * USD_TO_EUR, 2),
            "Account ID": "",
            "Loan Account ID": "",
            "Loan Status": "",
            "Status": "FAILED",
            "Reason": "",
            "Timestamp": datetime.now().isoformat(timespec="seconds"),
        }

        if errors:
            row["Reason"] = "; ".join(errors)
            log.warning(f"Skipping {username}: {row['Reason']}")
            report.append(row)
            continue

        driver = None
        try:
            driver = make_driver()
            driver.get(BASE_URL)
            register(driver, user)
            row["Account ID"] = open_account(driver)
            status, loan_acc = request_loan(driver, down_payment)
            row["Loan Status"] = status
            row["Loan Account ID"] = loan_acc
            row["Status"] = "COMPLETED"
            driver.find_element(By.LINK_TEXT, "Log Out").click()
        except Exception as e:
            row["Reason"] = str(e)[:200]
            if driver:
                driver.save_screenshot(SCREENSHOT_DIR / f"error_{username}.png")
        finally:
            if driver:
                driver.quit()

        report.append(row)

    df_report = pd.DataFrame(report)
    df_report.to_excel(REPORT_XLSX, index=False)
    format_excel_report(REPORT_XLSX, df_report)
    log.info(f" Report saved: {REPORT_XLSX}")


if __name__ == "__main__":
    main()
